import os
import shutil
import subprocess
import tempfile
from datetime import date, datetime, time

import openpyxl
from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from backend.models import Facility
from samples.models import SampleWithIssue


class Command(BaseCommand):
	help = "Import DBS, Plasma, and EID samples with missing details from the tracking spreadsheet."

	def add_arguments(self, parser):
		parser.add_argument('spreadsheet', help='Path to samples_without_enough_details.xls or a converted .xlsx file.')
		parser.add_argument('--database', default='default', help='Database alias to import into. Default: default.')
		parser.add_argument('--created-by', dest='created_by', help='Username to record as created_by for imported rows.')
		parser.add_argument('--test-type', choices=['VL', 'EID', 'HEP'], help='Optional test type for DBS and Plasma rows.')
		parser.add_argument('--commit', action='store_true', help='Save imported rows. Without this flag, only reports what would change.')

	def handle(self, *args, **options):
		path = options['spreadsheet']
		if not os.path.exists(path):
			raise CommandError('Spreadsheet not found: {0}'.format(path))

		created_by = self._created_by(options.get('created_by'), options['database'])
		workbook_path, cleanup_dir = self._workbook_path(path)
		try:
			stats = self._import_workbook(workbook_path, options['database'], created_by, options.get('test_type'), options['commit'])
		finally:
			if cleanup_dir:
				shutil.rmtree(cleanup_dir, ignore_errors=True)

		if not options['commit']:
			self.stdout.write(self.style.WARNING('Dry run only. Re-run with --commit to save changes.'))
		self.stdout.write(
			self.style.SUCCESS(
				'Processed {processed} row(s): {created} new, {updated} updated, {skipped} skipped.'.format(**stats)
			)
		)

	def _created_by(self, username, database):
		if not username:
			return None
		user = User.objects.using(database).filter(username=username).first()
		if not user:
			raise CommandError('User not found on {0}: {1}'.format(database, username))
		return user

	def _workbook_path(self, path):
		ext = os.path.splitext(path)[1].lower()
		if ext == '.xlsx':
			return path, None
		if ext != '.xls':
			raise CommandError('Expected .xls or .xlsx file.')
		tmp_dir = tempfile.mkdtemp(prefix='samples-with-issues-')
		cmd = ['libreoffice', '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, path]
		try:
			subprocess.check_call(cmd)
		except (OSError, subprocess.CalledProcessError) as exc:
			shutil.rmtree(tmp_dir, ignore_errors=True)
			raise CommandError('Could not convert .xls file. Install LibreOffice or provide .xlsx. {0}'.format(exc))
		converted = os.path.join(tmp_dir, os.path.splitext(os.path.basename(path))[0] + '.xlsx')
		if not os.path.exists(converted):
			shutil.rmtree(tmp_dir, ignore_errors=True)
			raise CommandError('LibreOffice did not produce the expected .xlsx file.')
		return converted, tmp_dir

	def _import_workbook(self, path, database, created_by, default_test_type, commit):
		stats = {'processed': 0, 'created': 0, 'updated': 0, 'skipped': 0}
		workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
		rows = []
		for sheet in workbook.worksheets:
			rows.extend(self._sheet_rows(sheet, database, created_by, default_test_type, stats))
		if commit and rows:
			with transaction.atomic(using=database):
				for row in rows:
					_, created = SampleWithIssue.objects.using(database).update_or_create(
						source_sheet=row.pop('source_sheet'),
						source_row=row.pop('source_row'),
						defaults=row,
					)
					if created:
						stats['created'] += 1
					else:
						stats['updated'] += 1
		else:
			stats['created'] = len(rows)
		return stats

	def _sheet_rows(self, sheet, database, created_by, default_test_type, stats):
		title = sheet.title.strip()
		header_row_number, headers = self._find_header(sheet)
		if not headers:
			stats['skipped'] += sheet.max_row
			return []
		out = []
		for row_number, row in enumerate(sheet.iter_rows(min_row=header_row_number + 1, values_only=True), start=header_row_number + 1):
			values = self._row_dict(headers, row)
			if not any(self._clean(value) for value in values.values()):
				stats['skipped'] += 1
				continue
			payload = self._payload(title, row_number, values, database, created_by, default_test_type)
			if not payload:
				stats['skipped'] += 1
				continue
			stats['processed'] += 1
			out.append(payload)
		return out

	def _find_header(self, sheet):
		for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
			headers = [self._normalize_header(cell) for cell in row]
			if 'reception date' in headers and ('barcode' in headers or 'infant name' in headers):
				return row_number, headers
		return None, None

	def _row_dict(self, headers, row):
		values = {}
		for index, header in enumerate(headers):
			if not header:
				continue
			value = row[index] if index < len(row) else None
			if header not in values or values[header] in [None, '']:
				values[header] = value
		return values

	def _payload(self, sheet_title, row_number, values, database, created_by, default_test_type):
		is_eid = 'eid' in sheet_title.lower()
		sample_type = None if is_eid else ('D' if 'dbs' in sheet_title.lower() else 'P')
		facility_name = self._clean(values.get('facility'))
		retrieved = self._yes(values.get('retrieval status'))
		return {
			'source_sheet': sheet_title,
			'source_row': row_number,
			'reception_date': self._date(values.get('reception date')),
			'pack_number': self._clean(values.get('pack number') or values.get('box/position') or values.get('ziplock')),
			'barcode': self._clean(values.get('barcode')),
			'facility_name': facility_name,
			'facility': self._facility(facility_name, database),
			'art_number': self._clean(values.get('art number') or values.get('art no')),
			'form_number': self._clean(values.get('form number') or values.get('form no')),
			'sample_type': sample_type,
			'test_type': 'EID' if is_eid else default_test_type,
			'collection_date': self._date(values.get('collection date')),
			'infant_name': self._clean(values.get('infant name')),
			'batch_number': self._clean(values.get('batch no')),
			'exp_number': self._clean(values.get('exp no')),
			'contact': self._clean(values.get('contact')),
			'retrieval_status': retrieved,
			'retrieval_date': self._datetime(values.get('retrieval date')),
			'initials': self._clean(values.get('initials') or values.get('ln')),
			'created_by': created_by,
		}

	def _facility(self, facility_name, database):
		if not facility_name or facility_name.lower() == 'none':
			return None
		return Facility.objects.using(database).filter(facility__iexact=facility_name).first()

	def _normalize_header(self, value):
		value = self._clean(value).lower().replace(':', '')
		value = ' '.join(value.split())
		mapping = {
			'facility ': 'facility',
			'art no': 'art no',
			'form no': 'form no',
			'batch no': 'batch no',
			'exp no': 'exp no',
			'ln': 'ln',
		}
		return mapping.get(value, value)

	def _clean(self, value):
		if value is None:
			return ''
		if isinstance(value, float) and value.is_integer():
			return str(int(value))
		if isinstance(value, int):
			return str(value)
		if isinstance(value, datetime):
			return value.strftime('%Y-%m-%d')
		if isinstance(value, date):
			return value.strftime('%Y-%m-%d')
		value = str(value).strip()
		return '' if value.lower() == 'nan' else value

	def _date(self, value):
		if isinstance(value, datetime):
			return value.date()
		if isinstance(value, date):
			return value
		value = self._clean(value)
		if not value:
			return None
		for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%b-%Y']:
			try:
				return datetime.strptime(value, fmt).date()
			except ValueError:
				pass
		return None

	def _datetime(self, value):
		if isinstance(value, datetime):
			return value
		parsed = self._date(value)
		return datetime.combine(parsed, time.min) if parsed else None

	def _yes(self, value):
		return self._clean(value).lower() in ['yes', 'y', '1', 'true', 'retrieved']
